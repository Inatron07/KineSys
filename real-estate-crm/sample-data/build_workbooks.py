"""
Builds sample Excel workbooks for the Real Estate CRM demo.
These represent the files that would live in the watched OneDrive folder —
Automation Anywhere reads them and upserts rows into the CRM via API.
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="808080")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INR_FMT = '"Rs. "#,##0'
DATE_FMT = "dd-mmm-yyyy"


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def style_body(ws, first_row, last_row, ncols, formats=None):
    formats = formats or {}
    for r in range(first_row, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c in formats:
                cell.number_format = formats[c]


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_note(ws, row, ncols, text):
    ws.cell(row=row, column=1, value=text).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


# ---------------------------------------------------------------------------
# 1. Leads_Import.xlsx
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Leads"

headers = ["Lead ID", "Date Captured", "Source", "Name", "Phone", "Email",
           "Project Interested", "Budget Min (Rs.)", "Budget Max (Rs.)",
           "Location Preference", "Assigned Broker", "Status",
           "Last Follow-up", "Next Follow-up", "Notes"]
ws.append(headers)

rows = [
    ["LD-1001", datetime.date(2026, 8, 1), "99acres", "Ramesh Nair", "+91 98200 11122", "ramesh.nair@example.com",
     "Skyline Meadows", 6500000, 8500000, "Baner, Pune", "Priya Kulkarni", "Site Visit Scheduled",
     datetime.date(2026, 8, 10), datetime.date(2026, 8, 16), "Prefers 2BHK, wants parking on same floor"],
    ["LD-1002", datetime.date(2026, 8, 2), "Facebook Ads", "Anjali Deshmukh", "+91 90210 33445", "anjali.d@example.com",
     "Orchid Residency", 9500000, 12000000, "Whitefield, Bangalore", "Arjun Mehta", "Contacted",
     datetime.date(2026, 8, 9), datetime.date(2026, 8, 14), "Relocating from Mumbai, needs school nearby"],
    ["LD-1003", datetime.date(2026, 8, 3), "Website", "Suresh Iyer", "+91 99870 55667", "suresh.iyer@example.com",
     "Palm Grove Villas", 25000000, 30000000, "ECR, Chennai", "Priya Kulkarni", "Negotiation",
     datetime.date(2026, 8, 11), datetime.date(2026, 8, 15), "Wants villa with private garden"],
    ["LD-1004", datetime.date(2026, 8, 4), "Housing.com", "Fatima Sheikh", "+91 91234 77889", "fatima.sheikh@example.com",
     "Skyline Meadows", 5000000, 6500000, "Baner, Pune", "Rohit Verma", "New",
     None, datetime.date(2026, 8, 13), "First-time buyer, needs home loan guidance"],
    ["LD-1005", datetime.date(2026, 8, 4), "Referral", "Vikram Singh", "+91 98765 12309", "vikram.singh@example.com",
     "Orchid Residency", 11000000, 13000000, "Whitefield, Bangalore", "Arjun Mehta", "Booked",
     datetime.date(2026, 8, 12), None, "Booking confirmed, unit OR-B-1204"],
    ["LD-1006", datetime.date(2026, 8, 5), "MagicBricks", "Neha Kapoor", "+91 90000 22334", "neha.kapoor@example.com",
     "Palm Grove Villas", 20000000, 24000000, "ECR, Chennai", "Rohit Verma", "Contacted",
     datetime.date(2026, 8, 10), datetime.date(2026, 8, 17), "Budget-sensitive, comparing 2 other projects"],
    ["LD-1007", datetime.date(2026, 8, 6), "Walk-in", "Karan Malhotra", "+91 99900 44556", "karan.m@example.com",
     "Skyline Meadows", 7000000, 9000000, "Baner, Pune", "Priya Kulkarni", "Site Visit Done",
     datetime.date(2026, 8, 12), datetime.date(2026, 8, 18), "Liked 3BHK corner unit, awaiting spouse approval"],
    ["LD-1008", datetime.date(2026, 8, 7), "99acres", "Divya Rao", "+91 98111 66778", "divya.rao@example.com",
     "Orchid Residency", 8500000, 10000000, "Whitefield, Bangalore", "Arjun Mehta", "New",
     None, datetime.date(2026, 8, 14), "Called once, phone unreachable second attempt"],
    ["LD-1009", datetime.date(2026, 8, 8), "Facebook Ads", "Mohammed Ali", "+91 97654 88990", "mohammed.ali@example.com",
     "Palm Grove Villas", 27000000, 32000000, "ECR, Chennai", "Rohit Verma", "Lost",
     datetime.date(2026, 8, 9), None, "Went with a competitor project, price sensitivity"],
    ["LD-1010", datetime.date(2026, 8, 9), "Website", "Sneha Joshi", "+91 96543 11220", "sneha.joshi@example.com",
     "Skyline Meadows", 6000000, 7500000, "Baner, Pune", "Priya Kulkarni", "New",
     None, datetime.date(2026, 8, 15), "Auto-captured from inbox, awaiting first contact"],
]
for r in rows:
    ws.append(r)

n = len(headers)
last = len(rows) + 1
style_header(ws, n)
style_body(ws, 2, last, n, formats={2: DATE_FMT, 8: INR_FMT, 9: INR_FMT, 13: DATE_FMT, 14: DATE_FMT})
autosize(ws, [10, 14, 12, 16, 16, 24, 18, 15, 15, 20, 16, 18, 13, 13, 34])

dv_status = DataValidation(type="list", formula1='"New,Contacted,Site Visit Scheduled,Site Visit Done,Negotiation,Booked,Lost"', allow_blank=True)
ws.add_data_validation(dv_status)
dv_status.add(f"L2:L{last}")

add_note(ws, last + 2, n,
         "Dropped into OneDrive by marketing/portal export. Automation Anywhere reads new/changed rows and "
         "upserts into the CRM via the leads API on Lead ID. Legend: edit any row; add new rows below the last one.")
wb.save("Leads_Import.xlsx")

# ---------------------------------------------------------------------------
# 2. Brokers.xlsx
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Brokers"

headers = ["Broker ID", "Name", "Phone", "Email", "Territory / Zone",
           "Active Leads", "Leads Assigned (MTD)", "Conversions (MTD)",
           "Conversion Rate", "Commission Rate", "Status"]
ws.append(headers)

data = [
    ["BR-201", "Priya Kulkarni", "+91 98220 10101", "priya.kulkarni@kinesys-crm.demo", "Baner, Pune", 4, 15, 3, None, 0.02, "Active"],
    ["BR-202", "Arjun Mehta", "+91 98230 20202", "arjun.mehta@kinesys-crm.demo", "Whitefield, Bangalore", 3, 12, 2, None, 0.02, "Active"],
    ["BR-203", "Rohit Verma", "+91 98240 30303", "rohit.verma@kinesys-crm.demo", "ECR, Chennai", 3, 11, 1, None, 0.015, "Active"],
    ["BR-204", "Sunita Rao", "+91 98250 40404", "sunita.rao@kinesys-crm.demo", "Hinjewadi, Pune", 0, 6, 1, None, 0.02, "Active"],
    ["BR-205", "Imran Qureshi", "+91 98260 50505", "imran.qureshi@kinesys-crm.demo", "Whitefield, Bangalore", 0, 4, 0, None, 0.015, "Inactive"],
]
for r in data:
    ws.append(r)

n = len(headers)
last = len(data) + 1
for r in range(2, last + 1):
    ws.cell(row=r, column=9, value=f"=IF(G{r}=0,0,H{r}/G{r})")

style_header(ws, n)
style_body(ws, 2, last, n, formats={9: "0.0%", 10: "0.0%"})
autosize(ws, [10, 16, 16, 30, 20, 12, 18, 16, 14, 14, 10])

dv_status = DataValidation(type="list", formula1='"Active,Inactive"', allow_blank=True)
ws.add_data_validation(dv_status)
dv_status.add(f"K2:K{last}")

add_note(ws, last + 2, n,
         "Conversion Rate is a formula (Conversions / Leads Assigned) — recalculates automatically. "
         "Active Leads count feeds the round-robin/workload assignment logic used by the inbox-capture bot.")
wb.save("Brokers.xlsx")

# ---------------------------------------------------------------------------
# 3. Inventory.xlsx
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Inventory"

headers = ["Project Name", "Unit No", "Tower / Block", "Unit Type",
           "Carpet Area (sqft)", "Price (Rs.)", "Status", "Booked By (Lead ID)",
           "Broker", "Booking Date"]
ws.append(headers)

data = [
    ["Skyline Meadows", "SM-A-101", "Tower A", "2BHK", 950, 6800000, "Available", None, None, None],
    ["Skyline Meadows", "SM-A-102", "Tower A", "2BHK", 950, 6800000, "Held", None, "Priya Kulkarni", None],
    ["Skyline Meadows", "SM-B-201", "Tower B", "3BHK", 1350, 9200000, "Available", None, None, None],
    ["Skyline Meadows", "SM-B-202", "Tower B", "3BHK", 1350, 9200000, "Booked", "LD-1007", "Priya Kulkarni", datetime.date(2026, 8, 13)],
    ["Orchid Residency", "OR-B-1204", "Block B", "3BHK", 1600, 11800000, "Sold", "LD-1005", "Arjun Mehta", datetime.date(2026, 8, 12)],
    ["Orchid Residency", "OR-A-905", "Block A", "2BHK", 1100, 9600000, "Available", None, None, None],
    ["Orchid Residency", "OR-A-906", "Block A", "2BHK", 1100, 9600000, "Available", None, None, None],
    ["Palm Grove Villas", "PGV-14", "Phase 1", "Villa - 4BHK", 3200, 27500000, "Held", "LD-1003", "Priya Kulkarni", None],
    ["Palm Grove Villas", "PGV-15", "Phase 1", "Villa - 4BHK", 3200, 27500000, "Available", None, None, None],
    ["Palm Grove Villas", "PGV-22", "Phase 2", "Villa - 3BHK", 2600, 21000000, "Available", None, None, None],
]
for r in data:
    ws.append(r)

n = len(headers)
last = len(data) + 1
style_header(ws, n)
style_body(ws, 2, last, n, formats={6: INR_FMT, 10: DATE_FMT})
autosize(ws, [18, 12, 14, 14, 16, 14, 12, 16, 16, 14])

dv_status = DataValidation(type="list", formula1='"Available,Held,Booked,Sold"', allow_blank=True)
ws.add_data_validation(dv_status)
dv_status.add(f"G2:G{last}")

add_note(ws, last + 2, n,
         "Sales ops maintains this in OneDrive. Status change to Booked/Sold is what the CRM watches to "
         "trigger the Bookings & Payments record and commission calculation.")
wb.save("Inventory.xlsx")

# ---------------------------------------------------------------------------
# 4. Bookings_Payments.xlsx
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Bookings_Payments"

headers = ["Booking ID", "Lead ID", "Client Name", "Unit", "Broker",
           "Total Value (Rs.)", "Amount Received (Rs.)", "Amount Due (Rs.)",
           "Next Payment Due", "Payment Status", "Commission Rate", "Commission Amount (Rs.)", "Commission Status"]
ws.append(headers)

data = [
    ["BK-501", "LD-1005", "Vikram Singh", "OR-B-1204", "Arjun Mehta", 11800000, 3540000, None,
     datetime.date(2026, 9, 5), "Partial", 0.02, None, "Pending"],
    ["BK-502", "LD-1007", "Karan Malhotra", "SM-B-202", "Priya Kulkarni", 9200000, 9200000, None,
     None, "Paid in Full", 0.02, None, "Pending"],
]
for r in data:
    ws.append(r)

n = len(headers)
last = len(data) + 1
for r in range(2, last + 1):
    ws.cell(row=r, column=8, value=f"=F{r}-G{r}")
    ws.cell(row=r, column=12, value=f"=F{r}*K{r}")

style_header(ws, n)
style_body(ws, 2, last, n, formats={6: INR_FMT, 7: INR_FMT, 8: INR_FMT, 9: DATE_FMT, 11: "0.0%", 12: INR_FMT})
autosize(ws, [10, 10, 18, 12, 16, 16, 18, 16, 16, 14, 13, 18, 15])

dv_pay = DataValidation(type="list", formula1='"Partial,Paid in Full,Overdue"', allow_blank=True)
ws.add_data_validation(dv_pay)
dv_pay.add(f"J2:J{last}")

dv_comm = DataValidation(type="list", formula1='"Pending,Paid"', allow_blank=True)
ws.add_data_validation(dv_comm)
dv_comm.add(f"M2:M{last}")

add_note(ws, last + 2, n,
         "Amount Due and Commission Amount are formulas — update Amount Received or Commission Rate and "
         "they recalculate. Accounts team reconciles this against bank statements via the RPA bot.")
wb.save("Bookings_Payments.xlsx")

print("Done: Leads_Import.xlsx, Brokers.xlsx, Inventory.xlsx, Bookings_Payments.xlsx")
